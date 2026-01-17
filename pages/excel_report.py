import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO

# -------------------------------------------------
# Appwrite backend
# -------------------------------------------------
from backend.appwrite_db import get_detailed_results


# -------------------------------------------------
# Normalize Appwrite data
# -------------------------------------------------
def normalize_data(raw_data):
    normalized = []

    for d in raw_data:
        normalized.append({
            "seat_no": str(d.get("Seat No", "")),
            "name": d.get("Name", ""),
            "percentage": d.get("Percentage", ""),
            "status": d.get("Status", ""),
            "code": d.get("Code", []) or [],
            "ua": d.get("UA", []) or [],
            "ca": d.get("CA", []) or [],
            "total": d.get("Total", []) or [],
            "status1": d.get("Status1", []) or []
        })

    return normalized


# -------------------------------------------------
# Excel generation
# -------------------------------------------------
# def create_excel_sheet():
#     raw_data = get_detailed_results()
#     data = normalize_data(raw_data)

#     if not data:
#         st.warning("No detailed student data available.")
#         return

#     with st.spinner("Creating Excel sheet..."):
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Student Results"

#         # -------------------------------------------------
#         # Collect all unique subject codes
#         # -------------------------------------------------
#         all_codes = []
#         for student in data:
#             for c in student["code"]:
#                 if isinstance(c, str) and c not in all_codes:
#                     all_codes.append(c)

#         # -------------------------------------------------
#         # Header
#         # -------------------------------------------------
#         header = ["Seat No", "Name"]
#         for code in all_codes:
#             header.extend([
#                 f"{code} UA",
#                 f"{code} CA",
#                 f"{code} Total",
#                 f"{code} Status"
#             ])
#         header.extend(["Grand Total", "Result", "Percentage"])
#         ws.append(header)

#         for cell in ws[1]:
#             cell.font = Font(bold=True)

#         # -------------------------------------------------
#         # Rows
#         # -------------------------------------------------
#         for student in data:
#             row = [student["seat_no"], student["name"]]

#             grand_total = 0
#             final_status = "Pass"

#             for code in all_codes:
#                 if code in student["code"]:
#                     i = student["code"].index(code)

#                     ua = student["ua"][i] if i < len(student["ua"]) else ""
#                     ca = student["ca"][i] if i < len(student["ca"]) else ""
#                     total = student["total"][i] if i < len(student["total"]) else ""
#                     status1 = student["status1"][i] if i < len(student["status1"]) else ""

#                     try:
#                         if str(total).isdigit():
#                             grand_total += int(total)
#                     except:
#                         pass

#                     if status1 != "P":
#                         final_status = "Fail"

#                     row.extend([ua, ca, total, status1])
#                 else:
#                     row.extend(["", "", "", ""])

#             # Percentage (assuming max 900)
#             try:
#                 percentage = f"{(grand_total / 900) * 100:.2f}"
#             except:
#                 percentage = "0.00"

#             row.extend([grand_total, final_status, percentage])
#             ws.append(row)

#         # -------------------------------------------------
#         # Download
#         # -------------------------------------------------
#         buffer = BytesIO()
#         wb.save(buffer)
#         buffer.seek(0)

#         st.success("Excel report generated successfully!")
#         st.download_button(
#             label="📥 Download Excel File",
#             data=buffer,
#             file_name="BCS_Results_Detailed.xlsx",
#             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )

def create_excel_sheet(data):
    if not data:
        st.warning("No student data available for selected filters.")
        return

    with st.spinner("Creating Excel sheet..."):
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Results"

        # -------------------------------------------------
        # Collect all unique subject codes
        # -------------------------------------------------
        all_codes = []
        for student in data:
            for c in student["code"]:
                if isinstance(c, str) and c not in all_codes:
                    all_codes.append(c)

        # -------------------------------------------------
        # Header
        # -------------------------------------------------
        header = ["Seat No", "Name"]
        for code in all_codes:
            header.extend([
                f"{code} UA",
                f"{code} CA",
                f"{code} Total",
                f"{code} Status"
            ])
        header.extend(["Grand Total", "Result", "Percentage"])
        ws.append(header)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        # -------------------------------------------------
        # Rows
        # -------------------------------------------------
        for student in data:
            row = [student["seat_no"], student["name"]]

            grand_total = 0
            final_status = "Pass"

            for code in all_codes:
                if code in student["code"]:
                    i = student["code"].index(code)

                    ua = student["ua"][i] if i < len(student["ua"]) else ""
                    ca = student["ca"][i] if i < len(student["ca"]) else ""
                    total = student["total"][i] if i < len(student["total"]) else ""
                    status1 = student["status1"][i] if i < len(student["status1"]) else ""

                    try:
                        if str(total).isdigit():
                            grand_total += int(total)
                    except:
                        pass

                    if status1 != "P":
                        final_status = "Fail"

                    row.extend([ua, ca, total, status1])
                else:
                    row.extend(["", "", "", ""])

            try:
                percentage = f"{(grand_total / 900) * 100:.2f}"
            except:
                percentage = "0.00"

            row.extend([grand_total, final_status, percentage])
            ws.append(row)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        st.success("Excel report generated successfully!")
        st.download_button(
            label="📥 Download Excel File",
            data=buffer,
            file_name="BCS_Results_Detailed.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


# -------------------------------------------------
# Streamlit entry
# -------------------------------------------------
# def show():
#     st.header("📝 Generate Detailed Excel Report")

#     # 🔽 REQUIRED FILTERS
#     course = st.selectbox("Course", ["BCS", "BCA"])
#     year = st.selectbox("Year", ["1", "2", "3"])
#     semester = st.selectbox("Semester", ["SEM-1", "SEM-3", "SEM-5"])
#     academic_year = st.text_input("Academic Year (e.g. 2025-26)")

#     if not academic_year:
#         st.info("Please enter Academic Year")
#         return

#     if st.button("Generate Excel"):
#         raw_data = get_detailed_results()

#         # 🔥 FILTER DATA FIRST
#         filtered_data = [
#             d for d in raw_data
#             if d["course"] == course
#             and d["year"] == year
#             and d["semester"] == semester
#             and d["academic_year"] == academic_year
#         ]

#         if not filtered_data:
#             st.warning("No records found for selected filters.")
#             return

#         create_excel_sheet(filtered_data)


def show():
    st.header("📝 Generate Detailed Excel Report")

    # -----------------------------
    # Session defaults
    # -----------------------------
    if "excel_continue" not in st.session_state:
        st.session_state.excel_continue = False

    if "excel_filtered" not in st.session_state:
        st.session_state.excel_filtered = []

    # -----------------------------
    # Filters (WITH KEYS)
    # -----------------------------
    course = st.selectbox(
        "Course", ["BCS", "BCA"], key="excel_course"
    )
    year = st.selectbox(
        "Year", ["1", "2", "3"], key="excel_year"
    )
    semester = st.selectbox(
        "Semester", ["SEM-1", "SEM-3", "SEM-5"], key="excel_sem"
    )
    academic_year = st.text_input(
        "Academic Year (e.g. 2025-26)", key="excel_ay"
    )

    col1, col2 = st.columns(2)

    # -----------------------------
    # CONTINUE
    # -----------------------------
    with col1:
        if st.button("➡️ Continue"):
            st.session_state.excel_continue = False
            st.session_state.excel_filtered = []

            if not academic_year:
                st.warning("Please enter Academic Year")
                return

            raw_data = get_detailed_results()
            if not raw_data:
                st.warning("No detailed data available.")
                return

            filtered = [
                d for d in raw_data
                if d["course"] == course
                and str(d["year"]) == str(year)
                and d["semester"] == semester
                and d["academic_year"] == academic_year
            ]

            if not filtered:
                st.warning("No records found for selected filters.")
                return

            st.session_state.excel_filtered = filtered
            st.session_state.excel_continue = True

    # -----------------------------
    # CLEAR
    # -----------------------------
    with col2:
        if st.button("🧹 Clear"):
            for k in [
                "excel_course",
                "excel_year",
                "excel_sem",
                "excel_ay",
                "excel_continue",
                "excel_filtered"
            ]:
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()

    # -----------------------------
    # SHOW SUMMARY + GENERATE
    # -----------------------------
    if st.session_state.excel_continue:
        st.divider()

        st.subheader("📌 Selected Result Details")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Course", course)
        c2.metric("Year", year)
        c3.metric("Semester", semester)
        c4.metric("Academic Year", academic_year)

        st.divider()

        if st.button("📥 Generate Excel"):
            create_excel_sheet(st.session_state.excel_filtered)



